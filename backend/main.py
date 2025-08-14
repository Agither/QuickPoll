from fastapi import FastAPI, HTTPException, Depends, Request, Response, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta
from enum import Enum
import uuid
import json
import random
import os
import tempfile
import asyncio
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Status Enum für Umfragen
class SurveyStatus(str, Enum):
    READY = "ready"  # Bereit zum Start
    ACTIVE = "active"  # Aktiv / Live
    FINISHED = "finished"  # Umfrage beendet

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# WebSocket Connection Manager
class WebSocketManager:
    def __init__(self):
        # survey_id -> {"hosts": [WebSocket], "participants": [WebSocket]}
        self.connections: Dict[str, Dict[str, List[WebSocket]]] = {}
        # websocket -> {"survey_id": str, "role": str, "session_id": str}
        self.connection_info: Dict[WebSocket, Dict[str, str]] = {}
        
    async def connect(self, websocket: WebSocket, survey_id: str, role: str, session_id: str):
        """WebSocket-Verbindung hinzufügen"""        
        if survey_id not in self.connections:
            self.connections[survey_id] = {"hosts": [], "participants": []}
            
        self.connections[survey_id][f"{role}s"].append(websocket)
        self.connection_info[websocket] = {
            "survey_id": survey_id,
            "role": role,
            "session_id": session_id
        }
        
        print(f"WebSocket connected: {role} for survey {survey_id}")
        
        # Initial Stats senden (ohne sofortiges Participant-Tracking)
        if role == "host":
            try:
                waiting_count = get_waiting_participants_count(survey_id)
                await websocket.send_text(json.dumps({
                    "type": "initial_stats",
                    "survey_id": survey_id,
                    "waiting_count": waiting_count
                }))
            except Exception as e:
                print(f"Error sending initial stats: {e}")
        
        # Bei Participant-Verbindung Tracking später in der WebSocket-Schleife
    
    async def disconnect(self, websocket: WebSocket):
        """WebSocket-Verbindung entfernen"""
        if websocket not in self.connection_info:
            return
            
        info = self.connection_info[websocket]
        survey_id = info["survey_id"]
        role = info["role"]
        session_id = info["session_id"]
        
        # Verbindung entfernen
        if survey_id in self.connections:
            role_list = self.connections[survey_id][f"{role}s"]
            if websocket in role_list:
                role_list.remove(websocket)
                
        del self.connection_info[websocket]
        
        # Bei Participant-Trennung aus Warteraum entfernen
        if role == "participant":
            track_participant_leave(survey_id, session_id)
            await self.broadcast_to_hosts(survey_id, {
                "type": "participant_left",
                "survey_id": survey_id,
                "session_id": session_id,
                "waiting_count": get_waiting_participants_count(survey_id)
            })
        
        print(f"WebSocket disconnected: {role} from survey {survey_id}")
    
    async def broadcast_to_hosts(self, survey_id: str, message: dict):
        """Nachricht an alle Hosts einer Umfrage senden"""
        if survey_id not in self.connections:
            return
            
        hosts = self.connections[survey_id]["hosts"].copy()
        
        dead_connections = []
        
        for host_ws in hosts:
            try:
                await host_ws.send_text(json.dumps(message))
            except:
                dead_connections.append(host_ws)
        
        # Tote Verbindungen entfernen
        for dead_ws in dead_connections:
            await self.disconnect(dead_ws)
    
    async def broadcast_to_participants(self, survey_id: str, message: dict):
        """Nachricht an alle Teilnehmer einer Umfrage senden"""
        if survey_id not in self.connections:
            return
            
        participants = self.connections[survey_id]["participants"].copy()
        dead_connections = []
        
        for participant_ws in participants:
            try:
                await participant_ws.send_text(json.dumps(message))
            except:
                dead_connections.append(participant_ws)
        
        # Tote Verbindungen entfernen
        for dead_ws in dead_connections:
            await self.disconnect(dead_ws)
    
    async def broadcast_to_all(self, survey_id: str, message: dict):
        """Nachricht an alle Verbindungen einer Umfrage senden"""
        await self.broadcast_to_hosts(survey_id, message)
        await self.broadcast_to_participants(survey_id, message)

# Globaler WebSocket Manager
ws_manager = WebSocketManager()

# In-Memory Teilnehmer-Tracking
participant_tracker = {}  # survey_id -> set of session_ids

def track_participant_join(survey_id: str, session_id: str):
    """Teilnehmer als 'wartend' markieren"""
    if survey_id not in participant_tracker:
        participant_tracker[survey_id] = set()
    participant_tracker[survey_id].add(session_id)
    print(f"Participant {session_id} joined survey {survey_id}. Total waiting: {len(participant_tracker[survey_id])}")

def track_participant_leave(survey_id: str, session_id: str):
    """Teilnehmer entfernen"""
    if survey_id in participant_tracker and session_id in participant_tracker[survey_id]:
        participant_tracker[survey_id].remove(session_id)
        if not participant_tracker[survey_id]:
            del participant_tracker[survey_id]
        print(f"Participant {session_id} left survey {survey_id}")

def get_waiting_participants_count(survey_id: str) -> int:
    """Anzahl wartender Teilnehmer abrufen"""
    return len(participant_tracker.get(survey_id, set()))

def get_submitted_participants_count(survey_id: str) -> int:
    """Anzahl der Teilnehmer die bereits abgestimmt haben"""
    with SessionLocal() as db:
        try:
            # Zähle eindeutige Responses für diese Umfrage
            from sqlalchemy import func, distinct
            count = db.query(func.count(distinct(ResponseDB.participant_name))).filter(
                ResponseDB.survey_id == survey_id
            ).scalar()
            return count or 0
        except Exception as e:
            logger.error(f"Error counting submitted participants: {e}")
            return 0

def get_survey_by_id(survey_id: str):
    """Hole Survey nach ID (einfache Version)"""
    with SessionLocal() as db:
        try:
            survey = db.query(SurveyDB).filter(SurveyDB.id == survey_id).first()
            if survey:
                return {
                    "id": survey.id,
                    "title": survey.title,
                    "description": survey.description,
                    "is_started": survey.is_started,
                    "created_at": survey.created_at
                }
            return None
        except Exception as e:
            logger.error(f"Error getting survey by ID: {e}")
            return None

def clear_waiting_participants(survey_id: str):
    """Alle wartenden Teilnehmer einer Umfrage entfernen"""
    if survey_id in participant_tracker:
        del participant_tracker[survey_id]
        print(f"Cleared all waiting participants for survey {survey_id}")

# SQLAlchemy Imports
from sqlalchemy import create_engine, String, DateTime, Boolean, Integer, Text, JSON, text, func, distinct
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, sessionmaker, Session

# Datenbank Setup für Railway mit Neon PostgreSQL
# Railway kann DATABASE_URL setzen, fallback zu Neon PostgreSQL
DATABASE_URL = os.getenv(
    "DATABASE_URL", 
    "postgresql://neondb_owner:npg_uIo6gw4kiTZa@ep-empty-frog-a2r2xjd6-pooler.eu-central-1.aws.neon.tech/neondb?sslmode=require"
)
logger.info(f"Using database: {DATABASE_URL}")

# SQLAlchemy Engine mit optimierten Einstellungen für Neon PostgreSQL
if "postgresql" in DATABASE_URL:
    # PostgreSQL für Neon/Railway Production
    engine = create_engine(
        DATABASE_URL,
        pool_pre_ping=True,  # Überprüft Verbindungen vor Verwendung
        pool_recycle=300,    # Erneuert Verbindungen alle 5 Minuten
        pool_size=10,        # Angepasst für Neon limits
        max_overflow=5,      # Zusätzliche Verbindungen bei Bedarf
        connect_args={"options": "-c timezone=utc"}  # UTC Timezone für konsistente Zeitstempel
    )
    logger.info("Configured Neon PostgreSQL database")
else:
    # SQLite für lokale Entwicklung
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
    logger.info("Configured SQLite database for local development")

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

class Base(DeclarativeBase):
    pass

# SQLAlchemy Models (Datenbank-Tabellen)
class SurveyDB(Base):
    __tablename__ = "surveys"
    
    id: Mapped[str] = mapped_column(String, primary_key=True, index=True)
    title: Mapped[str] = mapped_column(String, nullable=False)
    description: Mapped[str] = mapped_column(Text, nullable=True)
    status: Mapped[str] = mapped_column(String, default=SurveyStatus.READY.value)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.now)
    expires_at: Mapped[datetime] = mapped_column(DateTime, nullable=False)
    response_count: Mapped[int] = mapped_column(Integer, default=0)
    owner_session: Mapped[str] = mapped_column(String, nullable=True, index=True, default="")

class QuestionDB(Base):
    __tablename__ = "questions"
    
    id: Mapped[str] = mapped_column(String, primary_key=True, index=True)
    survey_id: Mapped[str] = mapped_column(String, nullable=False, index=True)
    title: Mapped[str] = mapped_column(String, nullable=False)
    type: Mapped[str] = mapped_column(String, nullable=False)
    options: Mapped[list] = mapped_column(JSON, nullable=True)  # Als JSON gespeichert
    required: Mapped[bool] = mapped_column(Boolean, default=True)
    description: Mapped[str] = mapped_column(Text, nullable=True)
    order: Mapped[int] = mapped_column(Integer, nullable=False)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.now)

class ResponseDB(Base):
    __tablename__ = "responses"
    
    id: Mapped[str] = mapped_column(String, primary_key=True, index=True)
    survey_id: Mapped[str] = mapped_column(String, nullable=False, index=True)
    participant_name: Mapped[str] = mapped_column(String, nullable=True)
    answers: Mapped[list] = mapped_column(JSON, nullable=False)  # Als JSON gespeichert
    submitted_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.now)

# Datenbank-Tabellen erstellen (mit Fehlerbehandlung für Railway)
try:
    logger.info("Creating database tables...")
    Base.metadata.create_all(bind=engine)
    logger.info("Database tables created successfully.")
except Exception as e:
    logger.warning(f"Database table creation failed (continuing anyway): {e}")
    # Fahre trotzdem fort - vielleicht sind die Tabellen bereits vorhanden

# Sanfte Migration für owner_session Feld
def ensure_owner_session_column():
    """Fügt owner_session Spalte hinzu falls sie nicht existiert (PostgreSQL-kompatibel)"""
    db = SessionLocal()
    try:
        logger.info("Prüfe owner_session Spalte...")
        
        # PostgreSQL: Prüfe Spalten über information_schema
        if "postgresql" in str(engine.url):
            result = db.execute(text("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'surveys' 
                AND column_name = 'owner_session'
            """)).fetchall()
            
            if not result:
                logger.info("Füge owner_session Spalte zur surveys Tabelle hinzu...")
                db.execute(text("ALTER TABLE surveys ADD COLUMN owner_session VARCHAR DEFAULT ''"))
                db.commit()
                logger.info("owner_session Spalte erfolgreich hinzugefügt.")
            else:
                logger.debug("owner_session Spalte bereits vorhanden.")
        else:
            # SQLite fallback für lokale Entwicklung
            result = db.execute(text("PRAGMA table_info(surveys)")).fetchall()
            columns = [row[1] for row in result]
            
            if 'owner_session' not in columns:
                logger.info("Füge owner_session Spalte zur surveys Tabelle hinzu...")
                db.execute(text("ALTER TABLE surveys ADD COLUMN owner_session TEXT DEFAULT ''"))
                db.commit()
                logger.info("owner_session Spalte erfolgreich hinzugefügt.")
            else:
                logger.debug("owner_session Spalte bereits vorhanden.")
                
    except Exception as e:
        logger.error(f"Migration Fehler: {e}")
        # Rollback bei Fehlern
        db.rollback()
    finally:
        db.close()

# Migration beim Start ausführen (mit Fehlerbehandlung für Railway)
try:
    logger.info("Running database migration...")
    ensure_owner_session_column()
    logger.info("Database migration completed.")
except Exception as e:
    logger.warning(f"Database migration failed (continuing anyway): {e}")
    # Fahre trotzdem fort - Migration ist optional

# Pydantic Models für API (Request/Response)
class QuestionType(str, Enum):
    MULTIPLE_CHOICE = "multiple_choice"
    SINGLE_CHOICE = "single_choice"
    TEXT = "text"
    RATING = "rating"
    YES_NO = "yes_no"

class QuestionBase(BaseModel):
    title: str = Field(..., min_length=1, max_length=500)
    type: QuestionType
    options: Optional[List[str]] = None
    required: bool = True
    description: Optional[str] = None

class QuestionCreate(QuestionBase):
    pass

class Question(QuestionBase):
    id: str
    survey_id: str
    order: int
    created_at: datetime

    class Config:
        from_attributes = True

class SurveyBase(BaseModel):
    title: str = Field(..., min_length=1, max_length=200)
    description: Optional[str] = Field(None, max_length=1000)
    status: SurveyStatus = SurveyStatus.READY

class SurveyCreate(SurveyBase):
    questions: List[QuestionCreate] = []

class Survey(SurveyBase):
    id: str
    created_at: datetime
    expires_at: datetime
    questions: List[Question] = []
    response_count: int = 0

    class Config:
        from_attributes = True

class AnswerSubmission(BaseModel):
    question_id: str
    answer: Any  # Kann String, List[str], int, bool sein

class ResponseSubmission(BaseModel):
    survey_id: str
    answers: List[AnswerSubmission]
    participant_name: Optional[str] = None

class Response(BaseModel):
    id: str
    survey_id: str
    answers: List[AnswerSubmission]
    participant_name: Optional[str]
    submitted_at: datetime

    class Config:
        from_attributes = True

# Dependency für Datenbankverbindung
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# FastAPI App initialisieren
app = FastAPI(
    title="QuickPoll API",
    description="QuickPoll API für Railway Deployment mit Neon PostgreSQL",
    version="2.0.0"
)

# Startup Event für Railway
@app.on_event("startup")
async def startup_event():
    logger.info("🚀 QuickPoll Backend starting on Railway...")
    logger.info(f"📊 Database: {DATABASE_URL}")
    logger.info(f"🔧 Health Check: /health")
    logger.info(f"📋 API Docs: /docs")
    
    # Teste Datenbank-Verbindung beim Start
    try:
        with SessionLocal() as db:
            db.execute(text("SELECT 1"))
        logger.info("✅ Database connection successful")
    except Exception as e:
        logger.warning(f"⚠️ Database connection failed: {e}")

@app.on_event("shutdown")
async def shutdown_event():
    logger.info("🛑 QuickPoll Backend shutting down...")

# CORS Middleware für Frontend-Integration
# Permissive CORS für Debugging
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Temporarily allow all origins
    allow_credentials=False,  # Must be False with allow_origins=["*"]
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

# Global Exception Handler für bessere Fehlerbehandlung
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    """Handle all unhandled exceptions"""
    import traceback
    error_details = traceback.format_exc()
    print(f"Unhandled exception: {str(exc)}")
    print(f"Traceback: {error_details}")
    
    return JSONResponse(
        status_code=500,
        content={"detail": f"Internal server error: {str(exc)}"},
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
            "Access-Control-Allow-Headers": "*",
        }
    )

# Helper Functions
def generate_session_id() -> str:
    """Generiert eine eindeutige Session-ID"""
    return str(uuid.uuid4()) + "-" + str(int(datetime.now().timestamp()))

def get_session_id_from_header(request: Request) -> str:
    """Extrahiert Session-ID aus Header, falls vorhanden"""
    session_id = request.headers.get("X-Session-ID")
    if not session_id:
        session_id = generate_session_id()
    return session_id

def check_survey_ownership(db: Session, survey_id: str, session_id: str) -> bool:
    """Prüft ob eine Umfrage dem aktuellen Session gehört"""
    survey = db.query(SurveyDB).filter(SurveyDB.id == survey_id).first()
    if not survey:
        return False
    # Falls owner_session leer/None ist, gehört Umfrage niemandem (legacy)
    if not survey.owner_session:
        return True
    return survey.owner_session == session_id

def generate_survey_id(db: Session) -> str:
    """Generiert eine eindeutige 4-stellige Umfrage-ID"""
    while True:
        # Generiere eine 4-stellige Zahl
        survey_id = str(random.randint(1000, 9999))
        
        # Prüfe ob die ID bereits existiert
        existing = db.query(SurveyDB).filter(SurveyDB.id == survey_id).first()
        if not existing:
            return survey_id

def generate_id() -> str:
    """Generiert eine UUID für andere Entitäten (Questions, Responses)"""
    return str(uuid.uuid4())

def cleanup_expired_surveys(db: Session):
    """Löscht abgelaufene Umfragen automatisch"""
    now = datetime.now()
    expired_surveys = db.query(SurveyDB).filter(SurveyDB.expires_at <= now).all()
    
    for survey in expired_surveys:
        # Lösche zuerst alle Fragen der Umfrage
        db.query(QuestionDB).filter(QuestionDB.survey_id == survey.id).delete()
        # Lösche alle Antworten der Umfrage
        db.query(ResponseDB).filter(ResponseDB.survey_id == survey.id).delete()
        # Lösche die Umfrage selbst
        db.delete(survey)
    
    if expired_surveys:
        db.commit()
        print(f"Gelöscht: {len(expired_surveys)} abgelaufene Umfragen")

def get_survey_with_questions(db: Session, survey_id: str) -> Survey:
    """Umfrage mit allen Fragen aus der Datenbank laden"""
    survey_db = db.query(SurveyDB).filter(SurveyDB.id == survey_id).first()
    if not survey_db:
        raise HTTPException(status_code=404, detail="Umfrage nicht gefunden")
    
    questions_db = db.query(QuestionDB).filter(QuestionDB.survey_id == survey_id).order_by(QuestionDB.order).all()
    
    questions = []
    for q_db in questions_db:
        question = Question(
            id=q_db.id,
            survey_id=q_db.survey_id,
            title=q_db.title,
            type=q_db.type,
            options=q_db.options,
            required=q_db.required,
            description=q_db.description,
            order=q_db.order,
            created_at=q_db.created_at
        )
        questions.append(question)
    
    return Survey(
        id=survey_db.id,
        title=survey_db.title,
        description=survey_db.description,
        status=SurveyStatus(survey_db.status),
        created_at=survey_db.created_at,
        expires_at=survey_db.expires_at,
        response_count=survey_db.response_count,
        questions=questions
    )

# Survey Endpoints
@app.post("/surveys/", response_model=Survey, tags=["Surveys"])
async def create_survey(survey_data: SurveyCreate, request: Request, db: Session = Depends(get_db)):
    """
    Erstellt eine neue Umfrage mit Fragen in der SQLite-Datenbank.
    
    - **title**: Titel der Umfrage (erforderlich)
    - **description**: Beschreibung der Umfrage (optional)
    - **questions**: Liste der Fragen (optional, können später hinzugefügt werden)
    
    Umfragen haben eine 4-stellige ID und laufen nach 7 Tagen ab.
    """
    # Session-ID aus Header extrahieren
    session_id = get_session_id_from_header(request)
    
    # Bereinige abgelaufene Umfragen vor der Erstellung einer neuen
    cleanup_expired_surveys(db)
    
    survey_id = generate_survey_id(db)
    now = datetime.now()
    expires_at = now + timedelta(days=7)
    
    # Umfrage in Datenbank speichern (standardmäßig "ready")
    survey_db = SurveyDB(
        id=survey_id,
        title=survey_data.title,
        description=survey_data.description,
        status=SurveyStatus.READY.value,  # Standardmäßig "ready"
        created_at=now,
        expires_at=expires_at,
        response_count=0,
        owner_session=session_id  # Session-ID des Erstellers
    )
    db.add(survey_db)
    
    # Fragen erstellen und speichern
    questions = []
    for i, q_data in enumerate(survey_data.questions):
        question_id = generate_id()
        question_db = QuestionDB(
            id=question_id,
            survey_id=survey_id,
            title=q_data.title,
            type=q_data.type.value,
            options=q_data.options,
            required=q_data.required,
            description=q_data.description,
            order=i,
            created_at=datetime.now()
        )
        db.add(question_db)
        
        question = Question(
            id=question_id,
            survey_id=survey_id,
            title=q_data.title,
            type=q_data.type,
            options=q_data.options,
            required=q_data.required,
            description=q_data.description,
            order=i,
            created_at=datetime.now()
        )
        questions.append(question)
    
    db.commit()
    
    return Survey(
        id=survey_id,
        title=survey_data.title,
        description=survey_data.description,
        status=SurveyStatus.READY,
        created_at=now,
        expires_at=expires_at,
        response_count=0,
        questions=questions
    )

@app.get("/surveys/", response_model=List[Survey], tags=["Surveys"])
async def get_all_surveys(request: Request, db: Session = Depends(get_db)):
    """Alle Umfragen des aktuellen Sessions aus der Datenbank abrufen (bereinigt automatisch abgelaufene)"""
    try:
        # Session-ID extrahieren
        session_id = get_session_id_from_header(request)
        print(f"Getting surveys for session: {session_id}")
        
        # Bereinige abgelaufene Umfragen vor der Abfrage
        cleanup_expired_surveys(db)
        
        # Nur eigene Umfragen abrufen (Session-basiert)
        surveys_db = db.query(SurveyDB).filter(SurveyDB.owner_session == session_id).all()
        print(f"Found {len(surveys_db)} surveys for session {session_id}")
        
        surveys = []
        for survey_db in surveys_db:
            survey = get_survey_with_questions(db, survey_db.id)
            surveys.append(survey)
        
        return surveys
    
    except Exception as e:
        print(f"Error in get_all_surveys: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.get("/surveys/{survey_id}", response_model=Survey, tags=["Surveys"])
async def get_survey(survey_id: str, request: Request, db: Session = Depends(get_db)):
    """Eine spezifische Umfrage aus der Datenbank abrufen (nur eigene Umfragen)"""
    # Session-ID extrahieren
    session_id = get_session_id_from_header(request)
    
    # Ownership prüfen
    if not check_survey_ownership(db, survey_id, session_id):
        raise HTTPException(
            status_code=403, 
            detail="Access denied. You can only access your own surveys."
        )
    
    return get_survey_with_questions(db, survey_id)

# Public Endpoints (für Teilnehmer)
@app.get("/public/surveys/{survey_id}", response_model=Survey, tags=["Public"])
async def get_public_survey(survey_id: str, db: Session = Depends(get_db)):
    """Öffentlicher Zugriff auf eine Umfrage für Teilnehmer"""
    try:
        logger.info(f"Public request for survey ID: {survey_id}")
        # Debug: Check if any surveys exist
        total_surveys = db.query(SurveyDB).count()
        logger.info(f"Total surveys in database: {total_surveys}")
        
        survey = get_survey_with_questions(db, survey_id)
        return survey
    except HTTPException as e:
        logger.error(f"HTTPException for survey {survey_id}: {e.detail}")
        raise HTTPException(status_code=404, detail="Umfrage nicht gefunden oder nicht verfügbar")
    except Exception as e:
        logger.error(f"Unexpected error for survey {survey_id}: {str(e)}")
        raise HTTPException(status_code=500, detail="Interner Serverfehler")

@app.put("/surveys/{survey_id}", response_model=Survey, tags=["Surveys"])
async def update_survey(survey_id: str, survey_data: SurveyBase, db: Session = Depends(get_db)):
    """Umfrage-Details in der Datenbank aktualisieren"""
    survey_db = db.query(SurveyDB).filter(SurveyDB.id == survey_id).first()
    if not survey_db:
        raise HTTPException(status_code=404, detail="Umfrage nicht gefunden")
    
    survey_db.title = survey_data.title
    survey_db.description = survey_data.description
    survey_db.status = survey_data.status.value
    
    db.commit()
    return get_survey_with_questions(db, survey_id)

@app.put("/surveys/{survey_id}/status", response_model=Survey, tags=["Surveys"])
async def update_survey_status(survey_id: str, status: SurveyStatus, request: Request, db: Session = Depends(get_db)):
    """Status einer Umfrage ändern (nur eigene Umfragen)"""
    # Session-ID extrahieren und Ownership prüfen
    session_id = get_session_id_from_header(request)
    if not check_survey_ownership(db, survey_id, session_id):
        raise HTTPException(
            status_code=403, 
            detail="Access denied. You can only modify your own surveys."
        )
    
    survey_db = db.query(SurveyDB).filter(SurveyDB.id == survey_id).first()
    if not survey_db:
        raise HTTPException(status_code=404, detail="Umfrage nicht gefunden")
    
    survey_db.status = status.value
    
    db.commit()
    return get_survey_with_questions(db, survey_id)

@app.delete("/surveys/{survey_id}", tags=["Surveys"])
async def delete_survey(survey_id: str, request: Request, db: Session = Depends(get_db)):
    """Umfrage und alle zugehörigen Daten aus der Datenbank löschen (nur eigene Umfragen)"""
    # Session-ID extrahieren und Ownership prüfen
    session_id = get_session_id_from_header(request)
    if not check_survey_ownership(db, survey_id, session_id):
        raise HTTPException(
            status_code=403, 
            detail="Access denied. You can only delete your own surveys."
        )
    
    survey_db = db.query(SurveyDB).filter(SurveyDB.id == survey_id).first()
    if not survey_db:
        raise HTTPException(status_code=404, detail="Umfrage nicht gefunden")
    
    # Zugehörige Fragen und Antworten löschen
    db.query(QuestionDB).filter(QuestionDB.survey_id == survey_id).delete()
    db.query(ResponseDB).filter(ResponseDB.survey_id == survey_id).delete()
    db.query(SurveyDB).filter(SurveyDB.id == survey_id).delete()
    
    db.commit()
    return {"message": "Umfrage erfolgreich gelöscht"}

# Question Endpoints
@app.post("/surveys/{survey_id}/questions/", response_model=Question, tags=["Questions"])
async def add_question(survey_id: str, question_data: QuestionCreate, db: Session = Depends(get_db)):
    """Neue Frage zu einer Umfrage in der Datenbank hinzufügen"""
    survey_db = db.query(SurveyDB).filter(SurveyDB.id == survey_id).first()
    if not survey_db:
        raise HTTPException(status_code=404, detail="Umfrage nicht gefunden")
    
    # Aktuelle Anzahl Fragen ermitteln für Order
    question_count = db.query(QuestionDB).filter(QuestionDB.survey_id == survey_id).count()
    
    question_id = generate_id()
    question_db = QuestionDB(
        id=question_id,
        survey_id=survey_id,
        title=question_data.title,
        type=question_data.type.value,
        options=question_data.options,
        required=question_data.required,
        description=question_data.description,
        order=question_count,
        created_at=datetime.now()
    )
    
    db.add(question_db)
    db.commit()
    
    return Question(
        id=question_id,
        survey_id=survey_id,
        title=question_data.title,
        type=question_data.type,
        options=question_data.options,
        required=question_data.required,
        description=question_data.description,
        order=question_count,
        created_at=datetime.now()
    )

@app.put("/surveys/{survey_id}/questions/{question_id}", response_model=Question, tags=["Questions"])
async def update_question(survey_id: str, question_id: str, question_data: QuestionBase, db: Session = Depends(get_db)):
    """Frage in der Datenbank aktualisieren"""
    question_db = db.query(QuestionDB).filter(
        QuestionDB.id == question_id,
        QuestionDB.survey_id == survey_id
    ).first()
    
    if not question_db:
        raise HTTPException(status_code=404, detail="Frage nicht gefunden")
    
    question_db.title = question_data.title
    question_db.type = question_data.type.value
    question_db.options = question_data.options
    question_db.required = question_data.required
    question_db.description = question_data.description
    
    db.commit()
    
    return Question(
        id=question_db.id,
        survey_id=question_db.survey_id,
        title=question_db.title,
        type=question_db.type,
        options=question_db.options,
        required=question_db.required,
        description=question_db.description,
        order=question_db.order,
        created_at=question_db.created_at
    )

@app.delete("/surveys/{survey_id}/questions/{question_id}", tags=["Questions"])
async def delete_question(survey_id: str, question_id: str, db: Session = Depends(get_db)):
    """Frage aus der Datenbank löschen"""
    question_db = db.query(QuestionDB).filter(
        QuestionDB.id == question_id,
        QuestionDB.survey_id == survey_id
    ).first()
    
    if not question_db:
        raise HTTPException(status_code=404, detail="Frage nicht gefunden")
    
    db.query(QuestionDB).filter(QuestionDB.id == question_id).delete()
    db.commit()
    
    return {"message": "Frage erfolgreich gelöscht"}

# Response Endpoints
@app.post("/responses/", response_model=Response, tags=["Responses"])
async def submit_response(response_data: ResponseSubmission, db: Session = Depends(get_db)):
    """
    Antwort auf eine Umfrage in der Datenbank speichern.
    
    - **survey_id**: ID der Umfrage
    - **answers**: Liste der Antworten mit question_id und answer
    - **participant_name**: Name des Teilnehmers (optional)
    """
    # Umfrage existiert?
    survey_db = db.query(SurveyDB).filter(SurveyDB.id == response_data.survey_id).first()
    if not survey_db:
        raise HTTPException(status_code=404, detail="Umfrage nicht gefunden")
    
    # Pflichtfragen validieren
    required_questions = db.query(QuestionDB).filter(
        QuestionDB.survey_id == response_data.survey_id,
        QuestionDB.required == True
    ).all()
    
    required_question_ids = {q.id for q in required_questions}
    answered_question_ids = {a.question_id for a in response_data.answers}
    
    missing_questions = required_question_ids - answered_question_ids
    if missing_questions:
        raise HTTPException(
            status_code=400,
            detail=f"Erforderliche Fragen nicht beantwortet: {missing_questions}"
        )
    
    # Antwort in Datenbank speichern
    response_id = generate_id()
    answers_json = [{"question_id": a.question_id, "answer": a.answer} for a in response_data.answers]
    
    response_db = ResponseDB(
        id=response_id,
        survey_id=response_data.survey_id,
        participant_name=response_data.participant_name,
        answers=answers_json,
        submitted_at=datetime.now()
    )
    
    db.add(response_db)
    
    # Response Count aktualisieren
    survey_db.response_count += 1
    
    db.commit()
    
    # Live-Update an Hosts senden
    await ws_manager.broadcast_to_hosts(response_data.survey_id, {
        "type": "response_submitted",
        "survey_id": response_data.survey_id,
        "response_count": survey_db.response_count,
        "participant_name": response_data.participant_name,
        "submitted_at": datetime.now().isoformat()
    })
    
    return Response(
        id=response_id,
        survey_id=response_data.survey_id,
        participant_name=response_data.participant_name,
        answers=response_data.answers,
        submitted_at=datetime.now()
    )

@app.get("/surveys/{survey_id}/responses/", response_model=List[Response], tags=["Responses"])
async def get_survey_responses(survey_id: str, db: Session = Depends(get_db)):
    """Alle Antworten zu einer Umfrage aus der Datenbank abrufen"""
    responses_db = db.query(ResponseDB).filter(ResponseDB.survey_id == survey_id).all()
    
    responses = []
    for r_db in responses_db:
        answers = [AnswerSubmission(question_id=a["question_id"], answer=a["answer"]) for a in r_db.answers]
        response = Response(
            id=r_db.id,
            survey_id=r_db.survey_id,
            participant_name=r_db.participant_name,
            answers=answers,
            submitted_at=r_db.submitted_at
        )
        responses.append(response)
    
    return responses

@app.get("/responses/{response_id}", response_model=Response, tags=["Responses"])
async def get_response(response_id: str, db: Session = Depends(get_db)):
    """Eine spezifische Antwort aus der Datenbank abrufen"""
    response_db = db.query(ResponseDB).filter(ResponseDB.id == response_id).first()
    if not response_db:
        raise HTTPException(status_code=404, detail="Antwort nicht gefunden")
    
    answers = [AnswerSubmission(question_id=a["question_id"], answer=a["answer"]) for a in response_db.answers]
    
    return Response(
        id=response_db.id,
        survey_id=response_db.survey_id,
        participant_name=response_db.participant_name,
        answers=answers,
        submitted_at=response_db.submitted_at
    )

# Analytics Endpoints
@app.get("/surveys/{survey_id}/analytics/", tags=["Analytics"])
async def get_survey_analytics(survey_id: str, db: Session = Depends(get_db)):
    """
    Grundlegende Analyse-Daten für eine Umfrage aus der Datenbank.
    Zeigt Antwortverteilung für Multiple-Choice-Fragen.
    """
    survey_db = db.query(SurveyDB).filter(SurveyDB.id == survey_id).first()
    if not survey_db:
        raise HTTPException(status_code=404, detail="Umfrage nicht gefunden")
    
    questions_db = db.query(QuestionDB).filter(QuestionDB.survey_id == survey_id).all()
    responses_db = db.query(ResponseDB).filter(ResponseDB.survey_id == survey_id).all()
    
    analytics = {
        "survey_id": survey_id,
        "total_responses": len(responses_db),
        "questions_analytics": {}
    }
    
    for question in questions_db:
        question_responses = []
        for response in responses_db:
            for answer_data in response.answers:
                if answer_data["question_id"] == question.id:
                    question_responses.append(answer_data["answer"])
        
        if question.type in ["multiple_choice", "single_choice"]:
            # Antwortverteilung für Choice-Fragen
            answer_counts = {}
            for answer in question_responses:
                if isinstance(answer, list):
                    for choice in answer:
                        answer_counts[choice] = answer_counts.get(choice, 0) + 1
                else:
                    answer_counts[answer] = answer_counts.get(answer, 0) + 1
            
            analytics["questions_analytics"][question.id] = {
                "question_title": question.title,
                "question_type": question.type,
                "answer_distribution": answer_counts,
                "total_answers": len(question_responses)
            }
        
        elif question.type == "rating":
            # Durchschnittsbewertung für Rating-Fragen
            ratings = [int(r) for r in question_responses if isinstance(r, (int, str)) and str(r).isdigit()]
            avg_rating = sum(ratings) / len(ratings) if ratings else 0
            
            analytics["questions_analytics"][question.id] = {
                "question_title": question.title,
                "question_type": question.type,
                "average_rating": round(avg_rating, 2),
                "total_ratings": len(ratings),
                "rating_distribution": {str(i): ratings.count(i) for i in range(1, 6)}
            }
    
    return analytics

# Export Endpoint
@app.get("/surveys/{survey_id}/export/", tags=["Export"])
async def export_survey_to_excel(survey_id: str, db: Session = Depends(get_db)):
    """Exportiert Umfrage-Ergebnisse als Excel-Datei"""
    
    # Umfrage finden
    survey = db.query(SurveyDB).filter(SurveyDB.id == survey_id).first()
    if not survey:
        raise HTTPException(status_code=404, detail="Umfrage nicht gefunden")
    
    # Fragen laden
    questions = db.query(QuestionDB).filter(QuestionDB.survey_id == survey_id).order_by(QuestionDB.order).all()
    if not questions:
        raise HTTPException(status_code=404, detail="Keine Fragen für diese Umfrage gefunden")
    
    # Antworten laden
    responses = db.query(ResponseDB).filter(ResponseDB.survey_id == survey_id).all()
    
    # Excel-Datei erstellen
    wb = Workbook()
    ws = wb.active
    ws.title = "Umfrage Ergebnisse"
    
    # Header-Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Umfrage-Informationen
    ws['A1'] = "Umfrage-Titel:"
    ws['B1'] = survey.title
    ws['A2'] = "Beschreibung:"
    ws['B2'] = survey.description or "Keine Beschreibung"
    ws['A3'] = "Status:"
    ws['B3'] = survey.status
    ws['A4'] = "Erstellt am:"
    ws['B4'] = survey.created_at.strftime("%d.%m.%Y %H:%M")
    ws['A5'] = "Antworten gesamt:"
    ws['B5'] = len(responses)
    
    # Style für Umfrage-Info
    for row in range(1, 6):
        ws[f'A{row}'].font = Font(bold=True)
    
    current_row = 7
    
    # Für jede Frage
    for question_idx, question in enumerate(questions, 1):
        # Frage-Header
        ws[f'A{current_row}'] = f"Frage {question_idx}: {question.title}"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = header_fill
        ws[f'A{current_row}'].alignment = header_alignment
        
        # Merge cells für Frage-Header
        ws.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1
        
        # Fragen-spezifische Daten verarbeiten
        question_responses = []
        for response in responses:
            for answer in response.answers:
                if answer.get('question_id') == question.id:
                    question_responses.append({
                        'participant': response.participant_name or f"Teilnehmer {response.id[:8]}",
                        'answer': answer.get('answer', ''),
                        'submitted_at': response.submitted_at
                    })
        
        if question.type in ['single_choice', 'multiple_choice', 'yes_no']:
            # Auswahl-Fragen: Zusammenfassung
            ws[f'A{current_row}'] = "Option"
            ws[f'B{current_row}'] = "Anzahl"
            ws[f'C{current_row}'] = "Prozent"
            
            # Header-Style für Spalten
            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].font = Font(bold=True)
                ws[f'{col}{current_row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            current_row += 1
            
            # Optionen zählen
            option_counts = {}
            total_responses = len(question_responses)
            
            for resp in question_responses:
                answer = resp['answer']
                if question.type == 'multiple_choice':
                    # Multiple Choice kann mehrere Antworten haben
                    if isinstance(answer, list):
                        for option in answer:
                            option_counts[option] = option_counts.get(option, 0) + 1
                    else:
                        option_counts[answer] = option_counts.get(answer, 0) + 1
                else:
                    option_counts[answer] = option_counts.get(answer, 0) + 1
            
            # Optionen auflisten
            for option, count in option_counts.items():
                percentage = (count / total_responses * 100) if total_responses > 0 else 0
                ws[f'A{current_row}'] = option
                ws[f'B{current_row}'] = count
                ws[f'C{current_row}'] = f"{percentage:.1f}%"
                current_row += 1
                
        elif question.type == 'rating':
            # Rating-Fragen: Statistiken
            ws[f'A{current_row}'] = "Bewertung"
            ws[f'B{current_row}'] = "Anzahl"
            ws[f'C{current_row}'] = "Prozent"
            
            # Header-Style
            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].font = Font(bold=True)
                ws[f'{col}{current_row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            current_row += 1
            
            # Bewertungen zählen
            ratings = []
            for resp in question_responses:
                try:
                    rating = int(resp['answer'])
                    if 1 <= rating <= 5:
                        ratings.append(rating)
                except (ValueError, TypeError):
                    continue
            
            # Rating-Verteilung
            for rating in range(1, 6):
                count = ratings.count(rating)
                percentage = (count / len(ratings) * 100) if ratings else 0
                ws[f'A{current_row}'] = f"{rating} Stern{'e' if rating != 1 else ''}"
                ws[f'B{current_row}'] = count
                ws[f'C{current_row}'] = f"{percentage:.1f}%"
                current_row += 1
            
            # Durchschnitt
            if ratings:
                avg_rating = sum(ratings) / len(ratings)
                ws[f'A{current_row}'] = "Durchschnitt:"
                ws[f'B{current_row}'] = f"{avg_rating:.2f}"
                ws[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                
        elif question.type == 'text':
            # Text-Fragen: Alle Antworten auflisten
            ws[f'A{current_row}'] = "Teilnehmer"
            ws[f'B{current_row}'] = "Antwort"
            ws[f'C{current_row}'] = "Zeitpunkt"
            
            # Header-Style
            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].font = Font(bold=True)
                ws[f'{col}{current_row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            current_row += 1
            
            # Antworten auflisten
            for resp in question_responses:
                ws[f'A{current_row}'] = resp['participant']
                ws[f'B{current_row}'] = resp['answer'] or "Keine Antwort"
                ws[f'C{current_row}'] = resp['submitted_at'].strftime("%d.%m.%Y %H:%M")
                current_row += 1
        
        current_row += 2  # Leerzeile zwischen Fragen
    
    # Spaltenbreite anpassen
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Temporäre Datei erstellen
    temp_dir = tempfile.gettempdir()
    filename = f"umfrage_{survey_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(temp_dir, filename)
    
    # Excel-Datei speichern
    wb.save(filepath)
    
    # Datei als Download zurückgeben
    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Health Check
@app.get("/health/", tags=["Health"])
async def health_check(db: Session = Depends(get_db)):
    """API Gesundheitsstatus mit Datenbankstatistiken"""
    surveys_count = db.query(SurveyDB).count()
    responses_count = db.query(ResponseDB).count()
    
    return {
        "status": "healthy",
        "timestamp": datetime.now(),
        "database": "SQLite",
        "surveys_count": surveys_count,
        "responses_count": responses_count
    }

@app.get("/debug/surveys/", tags=["Debug"])
async def debug_surveys(db: Session = Depends(get_db)):
    """Debug: Liste aller Umfrage-IDs"""
    surveys = db.query(SurveyDB.id, SurveyDB.title, SurveyDB.status).all()
    return {
        "total_count": len(surveys),
        "surveys": [{"id": s.id, "title": s.title, "status": s.status} for s in surveys]
    }

# Root Endpoint
@app.get("/", tags=["Root"])
async def root():
    """Willkommen bei der QuickPool API"""
    return {
        "message": "🎯 QuickPoll Backend läuft auf Railway!",
        "status": "healthy",
        "database": "Neon PostgreSQL" if "postgresql" in DATABASE_URL else "SQLite",
        "docs": "/docs",
        "redoc": "/redoc",
        "health": "/health",
        "version": "2.0.0 (Railway)",
        "timestamp": datetime.now().isoformat()
    }

# CORS Preflight Handler
@app.options("/{path:path}")
async def options_handler(path: str):
    """Handle CORS preflight requests"""
    return Response(
        status_code=200,
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
            "Access-Control-Allow-Headers": "*",
            "Access-Control-Max-Age": "86400",
        }
    )

# Health Check für Railway
@app.get("/health")
async def health_check():
    """Health check endpoint für Railway deployment"""
    try:
        # Basis Health Check - Server ist erreichbar
        return {
            "status": "healthy",
            "service": "QuickPoll Backend",
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return JSONResponse(
            status_code=503,
            content={"status": "unhealthy", "error": str(e)}
        )

# Erweiterte Health Check mit Datenbank-Test
@app.get("/health/detailed")
async def detailed_health_check():
    """Detaillierter Health check mit Datenbank-Test"""
    try:
        # Teste Datenbankverbindung
        with SessionLocal() as db:
            db.execute(text("SELECT 1"))
        
        return {
            "status": "healthy",
            "database": "connected",
            "service": "QuickPoll Backend",
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Detailed health check failed: {e}")
        return JSONResponse(
            status_code=503,
            content={
                "status": "unhealthy", 
                "database": "disconnected",
                "error": str(e),
                "timestamp": datetime.now().isoformat()
            }
        )

# WebSocket Endpunkte für Railway (persistente Server-Verbindungen)
@app.websocket("/ws/host/{survey_id}")
async def websocket_host_endpoint(websocket: WebSocket, survey_id: str):
    """WebSocket für Survey-Hosts (ManageScreen/ResultScreen)"""
    session_id = websocket.query_params.get("session_id", "")
    
    await websocket.accept()
    await ws_manager.connect(websocket, survey_id, "host", session_id)
    
    try:
        # Sende initial Stats
        try:
            waiting_count = get_waiting_participants_count(survey_id)
            await websocket.send_text(json.dumps({
                "type": "initial_stats",
                "survey_id": survey_id,
                "waiting_count": waiting_count
            }))
        except Exception as e:
            print(f"Error sending initial stats: {e}")
        
        # Keep connection alive
        while True:
            try:
                # Timeout für receive, damit Connection nicht blockiert
                data = await asyncio.wait_for(websocket.receive_text(), timeout=30.0)
                message = json.loads(data)
                
                # Host-Commands verarbeiten
                if message.get("type") == "start_survey":
                    await handle_start_survey(survey_id, websocket)
                elif message.get("type") == "end_survey":
                    await handle_end_survey(survey_id, websocket)
                elif message.get("type") == "ping":
                    await websocket.send_text(json.dumps({"type": "pong"}))
                    
            except asyncio.TimeoutError:
                # Sende Heartbeat
                try:
                    await websocket.send_text(json.dumps({"type": "heartbeat"}))
                except:
                    break
            except WebSocketDisconnect:
                break
            except Exception as e:
                print(f"WebSocket host error: {e}")
                break
                
    except WebSocketDisconnect:
        pass
    finally:
        await ws_manager.disconnect(websocket)

@app.websocket("/ws/participant/{survey_id}")
async def websocket_participant_endpoint(websocket: WebSocket, survey_id: str):
    """WebSocket für Survey-Teilnehmer (PollScreen)"""
    session_id = websocket.query_params.get("session_id", "")
    
    await websocket.accept()
    await ws_manager.connect(websocket, survey_id, "participant", session_id)
    
    try:
        # Participant als wartend markieren
        track_participant_join(survey_id, session_id)
        await ws_manager.broadcast_to_hosts(survey_id, {
            "type": "participant_joined",
            "survey_id": survey_id,
            "session_id": session_id,
            "waiting_count": get_waiting_participants_count(survey_id)
        })
        
        # Keep connection alive
        while True:
            try:
                # Timeout für receive, damit Connection nicht blockiert
                data = await asyncio.wait_for(websocket.receive_text(), timeout=30.0)
                message = json.loads(data)
                
                if message.get("type") == "ping":
                    await websocket.send_text(json.dumps({"type": "pong"}))
                    
            except asyncio.TimeoutError:
                # Sende Heartbeat
                try:
                    await websocket.send_text(json.dumps({"type": "heartbeat"}))
                except:
                    break
            except WebSocketDisconnect:
                break
            except Exception as e:
                print(f"WebSocket participant error: {e}")
                break
                
    except WebSocketDisconnect:
        pass
    finally:
        await ws_manager.disconnect(websocket)

# WebSocket Command Handlers
async def handle_start_survey(survey_id: str, host_websocket: WebSocket):
    """Survey starten - allen Teilnehmern Bescheid geben"""
    try:
        # Status in Datenbank ändern
        db = SessionLocal()
        survey_db = db.query(SurveyDB).filter(SurveyDB.id == survey_id).first()
        if survey_db:
            survey_db.status = SurveyStatus.ACTIVE.value
            db.commit()
            
            # Allen Teilnehmern Bescheid geben
            await ws_manager.broadcast_to_participants(survey_id, {
                "type": "survey_started",
                "survey_id": survey_id,
                "message": "Die Umfrage wurde gestartet!"
            })
            
            # Host bestätigen
            await host_websocket.send_text(json.dumps({
                "type": "survey_start_confirmed",
                "survey_id": survey_id
            }))
            
        db.close()
    except Exception as e:
        print(f"Error starting survey: {e}")

async def handle_end_survey(survey_id: str, host_websocket: WebSocket):
    """Survey beenden"""
    try:
        # Status in Datenbank ändern
        db = SessionLocal()
        survey_db = db.query(SurveyDB).filter(SurveyDB.id == survey_id).first()
        if survey_db:
            survey_db.status = SurveyStatus.FINISHED.value
            db.commit()
            
            # Allen Teilnehmern Bescheid geben
            await ws_manager.broadcast_to_participants(survey_id, {
                "type": "survey_finished",
                "survey_id": survey_id,
                "message": "Die Umfrage wurde beendet!"
            })
            
            # Host bestätigen
            await host_websocket.send_text(json.dumps({
                "type": "survey_end_confirmed",
                "survey_id": survey_id
            }))
            
        db.close()
    except Exception as e:
        print(f"Error ending survey: {e}")

# Railway deployment: uvicorn wird über Procfile gestartet
# Das Backend ist jetzt Railway-ready mit:
# - Neon PostgreSQL Datenbank
# - WebSocket Support
# - Health Check Endpoint
# - Robuste Fehlerbehandlung